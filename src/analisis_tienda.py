import pandas as pd
from sqlalchemy import create_engine
from sqlalchemy.exc import SQLAlchemyError
from openpyxl import load_workbook

# Ruta del archivo CSV
csv_path = 'data/sales_data_sample.csv'

# Leer archivo CSV
df = pd.read_csv(csv_path, encoding='unicode_escape')

# 2. Separar DataFrames
customers = df[['CUSTOMERNAME', 'PHONE', 'ADDRESSLINE1', 'ADDRESSLINE2', 'CITY',
                'STATE', 'POSTALCODE', 'COUNTRY', 'TERRITORY', 'CONTACTFIRSTNAME', 'CONTACTLASTNAME']].copy()

products = df[['PRODUCTCODE', 'PRODUCTLINE', 'MSRP']].copy()

orders = df[['ORDERNUMBER', 'ORDERDATE', 'STATUS', 'QTR_ID', 'MONTH_ID', 'YEAR_ID', 
             'DEALSIZE', 'CUSTOMERNAME']].copy()

order_items = df[['ORDERNUMBER', 'PRODUCTCODE', 'QUANTITYORDERED', 'PRICEEACH', 
                  'ORDERLINENUMBER', 'SALES']].copy()

# 3. Renombrar columnas al estilo snake_case
customers.rename(columns={
    'CUSTOMERNAME': 'customer_name',
    'PHONE': 'phone',
    'ADDRESSLINE1': 'address_line1',
    'ADDRESSLINE2': 'address_line2',
    'CITY': 'city',
    'STATE': 'state',
    'POSTALCODE': 'postal_code',
    'COUNTRY': 'country',
    'TERRITORY': 'territory',
    'CONTACTFIRSTNAME': 'contact_first_name',
    'CONTACTLASTNAME': 'contact_last_name'
}, inplace=True)

products.rename(columns={
    'PRODUCTCODE': 'product_code',
    'PRODUCTLINE': 'product_line',
    'MSRP': 'msrp'
}, inplace=True)

orders.rename(columns={
    'ORDERNUMBER': 'order_number',
    'ORDERDATE': 'order_date',
    'STATUS': 'status',
    'QTR_ID': 'qtr_id',
    'MONTH_ID': 'month_id',
    'YEAR_ID': 'year_id',
    'CUSTOMERNAME': 'customer_name',
    'DEALSIZE': 'deal_size'
}, inplace=True)

order_items.rename(columns={
    'ORDERNUMBER': 'order_number',
    'PRODUCTCODE': 'product_code',
    'QUANTITYORDERED': 'quantity_ordered',
    'PRICEEACH': 'price_each',
    'ORDERLINENUMBER': 'order_line_number',
    'SALES': 'sales'
}, inplace=True)

# 4. Convertir la fecha al formato correcto
orders['order_date'] = pd.to_datetime(orders['order_date'], format='%m/%d/%Y %H:%M')

# 5. Eliminar duplicados según claves
customers = customers.drop_duplicates(subset='customer_name').copy()
products = products.drop_duplicates(subset='product_code').copy()
orders = orders.drop_duplicates(subset='order_number').copy()
order_items = order_items.drop_duplicates(subset=['order_number', 'product_code']).copy()

# 6. Conectar con PostgreSQL y validar
try:
    # Datos de conexión (modifica con los tuyos)
    USER = 'postgres'
    PASSWORD = '********'
    HOST = 'localhost'
    PORT = '5432'
    DB = 'tienda_db'
    engine = create_engine(f'postgresql://{USER}:{PASSWORD}@{HOST}:{PORT}/{DB}')

    with engine.begin() as connection:
        customers.to_sql('customers', con=connection, if_exists='append', index=False)
        products.to_sql('products', con=connection, if_exists='append', index=False)
        orders.to_sql('orders', con=connection, if_exists='append', index=False)
        order_items.to_sql('order_items', con=connection, if_exists='append', index=False)

    print("✅ Todos los datos se insertaron correctamente.")

except SQLAlchemyError as e:
    print("❌ Ocurrió un error. Se canceló toda la inserción (rollback automático).")
    print(e)


# RESOLVIENDO INCOGNITAS

# PRODUCTOS MAS CAROS
# Consulta SQL
query = "SELECT * FROM products ORDER BY msrp DESC LIMIT 10"
df1 = pd.read_sql_query(query, engine)

# CLIENTES DE ESPAÑA
query = "SELECT * FROM customers WHERE country = 'Spain'"
df2 = pd.read_sql_query(query, engine)

# Total de ventas por país
query = """
SELECT country, COUNT(order_number) AS total_pedidos FROM customers 
INNER JOIN orders on customers.customer_name = orders.customer_name GROUP BY country
"""
df3 = pd.read_sql_query(query, engine)

# ¿Cuántos productos se han vendido por categoría?
query = """
SELECT product_line, COUNT(product_line) AS 
cantidad_categoria FROM order_items INNER JOIN products ON order_items.product_code=products.product_code 
GROUP BY product_line
"""
df4 = pd.read_sql_query(query, engine)

# ¿Quiénes son los mejores clientes?
query = """
SELECT customer_name, COUNT(order_number) AS mejores_clientes 
FROM orders GROUP BY customer_name ORDER BY mejores_clientes DESC
"""
df5 = pd.read_sql_query(query, engine)

# Ventas por mes
query = """
SELECT TO_CHAR(order_date,'YYYY-MM') AS año_mes, COUNT(order_number) AS total_pedidos 
FROM customers INNER JOIN orders on customers.customer_name = orders.customer_name GROUP BY año_mes 
ORDER BY año_mes
"""
df6 = pd.read_sql_query(query, engine)

# ¿Cuál fue el mes con más ingresos?
query = """
SELECT TO_CHAR(order_date,'YYYY-MM') AS año_mes, COUNT(order_number) 
AS total_pedidos FROM customers INNER JOIN orders on customers.customer_name = orders.customer_name 
GROUP BY año_mes ORDER BY total_pedidos desc LIMIT 1
"""
df7 = pd.read_sql_query(query, engine)

# ¿Qué categoría generó más ingresos?
query = """
SELECT product_line, SUM(msrp) AS ingreso_total FROM order_items INNER JOIN products ON 
order_items.product_code=products.product_code GROUP BY product_line ORDER BY ingreso_total DESC
LIMIT 1
"""
df8 = pd.read_sql_query(query, engine)

# Lista de pedidos con nombre de cliente y país
query = """
SELECT customers.customer_name, country FROM customers 
INNER JOIN orders on customers.customer_name = orders.customer_name 
ORDER BY customers.customer_name
"""
df9 = pd.read_sql_query(query, engine)

#  Lista de productos comprados por cada cliente
query = """
SELECT customer_name, product_code FROM orders
INNER JOIN order_items ON orders.order_number=order_items.order_number
ORDER BY customer_name
"""
df10 = pd.read_sql_query(query, engine)

# Clientes que han comprado más que la media
query = """
SELECT customer_name, COUNT(order_number) AS total_pedidos_cliente
FROM orders
GROUP BY customer_name
HAVING 
	COUNT(order_number) > (
		SELECT AVG(cantidad_pedidos_clientes)
			FROM (
				SELECT customer_name, COUNT(order_number) AS cantidad_pedidos_clientes
				FROM orders
				GROUP BY customer_name
			)
	)
"""
df11 = pd.read_sql_query(query, engine)

# Productos que nunca se han vendido
query = """
SELECT * FROM products WHERE product_code NOT IN (SELECT product_code FROM order_items)
"""
df12 = pd.read_sql_query(query, engine)


# EXPORTAMOS LOS DATOS OBTENIDOS A UN ARCHIVO DE EXCEL CON VARIAS HOJAS

# Lista de preguntas (una para cada DataFrame)
preguntas = [
    "Pregunta 1: ¿Cuales son los productos mas caros?",
    "Pregunta 2: ¿Qué clientes son de españa?",
    "Pregunta 3: ¿Cuáles fueron los ingresos por paises?",
    "Pregunta 4: ¿Cuántos productos se han vendido por categoría?",
    "Pregunta 5: ¿Quiénes son los mejores clientes?",
    "Pregunta 6: ¿Cuales fueron las ventas por meses?",
    "Pregunta 7: ¿Cuál fue el mes con más ingresos??",
    "Pregunta 8: ¿Qué categoría generó más ingresos?",
    "Pregunta 9: Lista de pedidos con nombre de cliente y país",
    "Pregunta 10: Lista de productos comprados por cada cliente",
    "Pregunta 11: ¿Que clientes han comprado más que la media?",
    "Pregunta 12: ¿Cuáles son los productos que nunca se han vendido?"
]
# Lista de DataFrames
dfs = [df1, df2, df3, df4, df5, df6, df7, df8, df9, df10, df11, df12]

# Ruta del archivo Excel
archivo_excel = "data/reporte_completo.xlsx"

# Escribir DataFrames empezando desde la fila 2 (startrow=1)
with pd.ExcelWriter(archivo_excel, engine='openpyxl') as writer:
    for i, (df, pregunta) in enumerate(zip(dfs, preguntas), 1):
        hoja = f"Hoja{i}"
        df.to_excel(writer, sheet_name=hoja, startrow=1, index=False)
    

# Reabrir para insertar preguntas en la fila 1
wb = load_workbook(archivo_excel)

for i, pregunta in enumerate(preguntas, 1):
    hoja = f"Hoja{i}"
    ws = wb[hoja]
    ws.insert_rows(1)
    ws['A1'] = pregunta

wb.save(archivo_excel)